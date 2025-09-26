from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from sqlalchemy import text
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
import traceback
from dateutil.relativedelta import relativedelta

dashboard_bp = Blueprint('dashboard', __name__)

def decimal_to_float(obj):
    """Converter Decimal para float para serialização JSON"""
    if isinstance(obj, Decimal):
        return float(obj)
    return obj

def calculate_previous_month(date_filter):
    """
    Calcula o mês anterior baseado no formato YYYY-MM
    Retorna string no formato MM/YYYY para uso nas queries
    """
    try:
        year, month = date_filter.split('-')
        current_date = datetime(int(year), int(month), 1)
        previous_date = current_date - relativedelta(months=1)
        return previous_date.strftime('%m/%Y')
    except:
        return None

def calc_variacao(atual, anterior):
    """Calcula variação percentual entre dois valores"""
    if anterior and anterior > 0:
        return round(((atual - anterior) / anterior) * 100, 2)
    return None

def get_current_day_of_month():
    """Retorna o dia atual do mês como inteiro"""
    return datetime.now().day

# ========================================
# DASHBOARD COM ESTRUTURA DE FILTROS
# PRONTO PARA RECEBER CARDS COM QUERIES
# ========================================

@dashboard_bp.route('/api/dashboard/available-dates', methods=['GET'])
@jwt_required()
def get_available_dates():
    """
    Retorna as datas disponíveis na base para o filtro
    """
    try:
        with db.engine.connect() as conn:
            result = conn.execute(text("""
                SELECT to_char("data cadastro", 'MM/YYYY') AS mes_ano,
                       to_char("data cadastro", 'YYYY-MM') AS mes_valor,
                       MIN("data cadastro") AS data_ordenacao
                FROM public."V_CUSTOMER"
                WHERE "data cadastro" IS NOT NULL
                GROUP BY to_char("data cadastro", 'MM/YYYY'), to_char("data cadastro", 'YYYY-MM')
                ORDER BY MIN("data cadastro") DESC
            """))
            
            dates = []
            for row in result:
                dates.append({
                    'label': row[0],  # MM/YYYY format
                    'value': row[1]   # YYYY-MM format
                })
            
            return jsonify(dates), 200
            
    except Exception as e:
        print(f"Erro em get_available_dates: {str(e)}")
        return jsonify({'error': 'Erro ao buscar datas disponíveis'}), 500

@dashboard_bp.route('/api/dashboard/card/<card_name>', methods=['GET'])
@jwt_required()
def get_card_data(card_name):
    """
    Endpoint genérico para dados de cards
    Suporta filtro de data e modo consolidado
    Inclui cálculo de variação proporcional em relação ao mês anterior
    """
    try:
        # Parâmetros da requisição
        date_filter = request.args.get('date')  # YYYY-MM format
        consolidated = request.args.get('consolidated', 'false').lower() == 'true'
        
        with db.engine.connect() as conn:
            result = {'value': 0, 'available': False, 'change': None}
            
            # CARD: Total de Cadastrados
            if card_name == 'clientes_cadastrados':
                if consolidated:
                    # Query do número consolidado (sem variação)
                    query_result = conn.execute(text("""
                        SELECT COUNT(*) AS total_clientes_cadastrados_consolidados
                        FROM public."V_CUSTOMER"
                        WHERE "data cadastro" IS NOT NULL
                    """))
                    
                    if query_result:
                        row = query_result.fetchone()
                        result = {
                            'value': row[0] if row else 0,
                            'available': True,
                            'change': None  # Sem variação no modo consolidado
                        }
                else:
                    # Converter YYYY-MM para MM/YYYY
                    if date_filter:
                        year, month = date_filter.split('-')
                        date_format = f"{month}/{year}"
                        
                        # Query do número por mês ATUAL
                        current_query = conn.execute(text("""
                            SELECT COUNT(*) AS total_clientes_cadastrados_no_mes
                            FROM public."V_CUSTOMER"
                            WHERE to_char("data cadastro", 'MM/YYYY') = :date_format
                        """), {'date_format': date_format})
                        
                        current_row = current_query.fetchone()
                        current_value = current_row[0] if current_row else 0
                        
                        # Detectar se é o mês atual para decidir tipo de comparação
                        current_month_str = datetime.now().strftime('%m/%Y')
                        is_current_month = (date_format == current_month_str)
                        
                        # Calcular mês anterior e buscar dados para comparação
                        previous_month = calculate_previous_month(date_filter)
                        change_percentage = None
                        
                        if previous_month:
                            try:
                                if is_current_month:
                                    # Mês atual: comparação proporcional (até hoje vs mesmo dia do mês anterior)
                                    current_day = get_current_day_of_month()
                                    
                                    # Query proporcional do mês ANTERIOR até mesmo dia
                                    previous_query = conn.execute(text("""
                                        SELECT COUNT(*) AS total_clientes_cadastrados_mes_anterior
                                        FROM public."V_CUSTOMER"
                                        WHERE to_char("data cadastro", 'MM/YYYY') = :previous_date
                                        AND "data cadastro" IS NOT NULL
                                        AND EXTRACT(day FROM "data cadastro") <= :current_day
                                    """), {
                                        'previous_date': previous_month,
                                        'current_day': current_day
                                    })
                                    
                                    previous_row = previous_query.fetchone()
                                    previous_value = previous_row[0] if previous_row else 0
                                    
                                    # Query proporcional do mês ATUAL até hoje
                                    current_proportional_query = conn.execute(text("""
                                        SELECT COUNT(*) AS total_clientes_cadastrados_atual_proporcional
                                        FROM public."V_CUSTOMER"
                                        WHERE to_char("data cadastro", 'MM/YYYY') = :current_date
                                        AND "data cadastro" IS NOT NULL
                                        AND EXTRACT(day FROM "data cadastro") <= :current_day
                                    """), {
                                        'current_date': date_format,
                                        'current_day': current_day
                                    })
                                    
                                    current_proportional_row = current_proportional_query.fetchone()
                                    current_proportional_value = current_proportional_row[0] if current_proportional_row else 0
                                    
                                    # Usar o valor proporcional que acabamos de calcular
                                    change_percentage = calc_variacao(current_proportional_value, previous_value)
                                    
                                else:
                                    # Mês anterior: comparação completa (mês fechado vs mês fechado anterior)
                                    # Query completa do mês ANTERIOR
                                    previous_query = conn.execute(text("""
                                        SELECT COUNT(*) AS total_clientes_cadastrados_mes_anterior
                                        FROM public."V_CUSTOMER"
                                        WHERE to_char("data cadastro", 'MM/YYYY') = :previous_date
                                        AND "data cadastro" IS NOT NULL
                                    """), {
                                        'previous_date': previous_month
                                    })
                                    
                                    previous_row = previous_query.fetchone()
                                    previous_value = previous_row[0] if previous_row else 0
                                    
                                    # Usar o valor total do mês atual (já calculado acima)
                                    # Para mês fechado: comparar total vs total
                                    change_percentage = calc_variacao(current_value, previous_value)
                                
                            except Exception as e:
                                pass
                        
                        result = {
                            'value': current_value,
                            'available': True,
                            'change': change_percentage
                        }
                    else:
                        result = {'value': 0, 'available': False, 'change': None}
            
            # Card: Total de Ativações
            elif card_name == 'total_ativacoes':
                if consolidated:
                    # Query do número consolidado (sem variação)
                    query_result = conn.execute(text("""
                        SELECT COUNT(*) AS total_clientes_consolidados
                        FROM public."V_CUSTOMER"
                        WHERE "data ativo" IS NOT NULL
                    """))
                    
                    if query_result:
                        row = query_result.fetchone()
                        result = {
                            'value': row[0] if row else 0,
                            'available': True,
                            'change': None  # Sem variação no modo consolidado
                        }
                else:
                    # Converter YYYY-MM para MM/YYYY
                    if date_filter:
                        year, month = date_filter.split('-')
                        date_format = f"{month}/{year}"
                        
                        # Query do número por mês ATUAL
                        current_query = conn.execute(text("""
                            SELECT COUNT(*) AS total_clientes_no_mes
                            FROM public."V_CUSTOMER"
                            WHERE to_char("data ativo", 'MM/YYYY') = :date_format
                        """), {'date_format': date_format})
                        
                        current_row = current_query.fetchone()
                        current_value = current_row[0] if current_row else 0
                        
                        # Detectar se é o mês atual para decidir tipo de comparação
                        current_month_str = datetime.now().strftime('%m/%Y')
                        is_current_month = (date_format == current_month_str)
                        
                        # Calcular mês anterior e buscar dados para comparação
                        previous_month = calculate_previous_month(date_filter)
                        change_percentage = None
                        
                        if previous_month:
                            try:
                                if is_current_month:
                                    # Mês atual: comparação proporcional (até hoje vs mesmo dia do mês anterior)
                                    current_day = get_current_day_of_month()
                                    
                                    # Query proporcional do mês ANTERIOR até mesmo dia
                                    previous_query = conn.execute(text("""
                                        SELECT COUNT(*) AS total_clientes_mes_anterior
                                        FROM public."V_CUSTOMER"
                                        WHERE to_char("data ativo", 'MM/YYYY') = :previous_date
                                        AND "data ativo" IS NOT NULL
                                        AND EXTRACT(day FROM "data ativo") <= :current_day
                                    """), {
                                        'previous_date': previous_month,
                                        'current_day': current_day
                                    })
                                    
                                    previous_row = previous_query.fetchone()
                                    previous_value = previous_row[0] if previous_row else 0
                                    
                                    # Query proporcional do mês ATUAL até hoje
                                    current_proportional_query = conn.execute(text("""
                                        SELECT COUNT(*) AS total_clientes_atual_proporcional
                                        FROM public."V_CUSTOMER"
                                        WHERE to_char("data ativo", 'MM/YYYY') = :current_date
                                        AND "data ativo" IS NOT NULL
                                        AND EXTRACT(day FROM "data ativo") <= :current_day
                                    """), {
                                        'current_date': date_format,
                                        'current_day': current_day
                                    })
                                    
                                    current_proportional_row = current_proportional_query.fetchone()
                                    current_proportional_value = current_proportional_row[0] if current_proportional_row else 0
                                    
                                    # Usar o valor proporcional que acabamos de calcular
                                    change_percentage = calc_variacao(current_proportional_value, previous_value)
                                    
                                else:
                                    # Mês anterior: comparação completa (mês fechado vs mês fechado anterior)
                                    # Query completa do mês ANTERIOR
                                    previous_query = conn.execute(text("""
                                        SELECT COUNT(*) AS total_clientes_mes_anterior
                                        FROM public."V_CUSTOMER"
                                        WHERE to_char("data ativo", 'MM/YYYY') = :previous_date
                                        AND "data ativo" IS NOT NULL
                                    """), {
                                        'previous_date': previous_month
                                    })
                                    
                                    previous_row = previous_query.fetchone()
                                    previous_value = previous_row[0] if previous_row else 0
                                    
                                    # Usar o valor total do mês atual (já calculado acima)
                                    # Para mês fechado: comparar total vs total
                                    change_percentage = calc_variacao(current_value, previous_value)
                                
                            except Exception as e:
                                pass
                        
                        result = {
                            'value': current_value,
                            'available': True,
                            'change': change_percentage
                        }
                    else:
                        result = {'value': 0, 'available': False, 'change': None}
            
            return jsonify(result), 200
            
    except Exception as e:
        print(f"Erro em get_card_data para {card_name}: {str(e)}")
        return jsonify({'error': f'Erro ao buscar dados do card {card_name}'}), 500

def safe_excel_value(value):
    """Converter valor para formato seguro para Excel"""
    if value is None:
        return ''
    elif isinstance(value, Decimal):
        return float(value)
    elif isinstance(value, (datetime, )):
        # Datas datetime para string
        return value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
    elif hasattr(value, 'date'):
        # Objetos date para string
        return value.strftime('%Y-%m-%d')
    else:
        # Qualquer outro tipo: converter para string
        return str(value)

@dashboard_bp.route('/api/dashboard/export/<card_name>', methods=['GET'])
@jwt_required()
def export_card_data(card_name):
    """
    Endpoint para exportar listas detalhadas em Excel ou retornar dados para prévia
    """
    try:
        # Parâmetros da requisição
        date_filter = request.args.get('date')
        consolidated = request.args.get('consolidated', 'false').lower() == 'true'
        preview = request.args.get('preview', 'false').lower() == 'true'
        
        with db.engine.connect() as conn:
            query_result = None
            
            # CARD: Total de Cadastrados
            if card_name == 'clientes_cadastrados':
                if consolidated:
                    # Query com TODAS as colunas
                    query = """
                        SELECT *
                        FROM public."V_CUSTOMER"
                        WHERE "data cadastro" IS NOT NULL
                        ORDER BY "código" ASC
                    """
                    
                    # Adicionar LIMIT apenas para prévia
                    if preview:
                        query += " LIMIT 50"
                    
                    query_result = conn.execute(text(query))
                    
                else:
                    if date_filter:
                        year, month = date_filter.split('-')
                        date_format = f"{month}/{year}"
                        
                        # Query com TODAS as colunas
                        query = """
                            SELECT *
                            FROM public."V_CUSTOMER"
                            WHERE to_char("data cadastro", 'MM/YYYY') = :date_format
                            ORDER BY "código" ASC
                        """
                        
                        # Adicionar LIMIT apenas para prévia
                        if preview:
                            query += " LIMIT 50"
                        
                        query_result = conn.execute(text(query), {'date_format': date_format})
                    else:
                        return jsonify({'error': 'Filtro de data é obrigatório para consulta não consolidada'}), 400
            
            # Card: Total de Ativações
            elif card_name == 'total_ativacoes':
                if consolidated:
                    # Query com TODAS as colunas
                    query = """
                        SELECT *
                        FROM public."V_CUSTOMER"
                        WHERE "data ativo" IS NOT NULL
                        ORDER BY "código" ASC
                    """
                    
                    # Adicionar LIMIT apenas para prévia
                    if preview:
                        query += " LIMIT 50"
                    
                    query_result = conn.execute(text(query))
                else:
                    if date_filter:
                        year, month = date_filter.split('-')
                        date_format = f"{month}/{year}"
                        
                        # Query com TODAS as colunas
                        query = """
                            SELECT *
                            FROM public."V_CUSTOMER"
                            WHERE to_char("data ativo", 'MM/YYYY') = :date_format
                            ORDER BY "código" ASC
                        """
                        
                        # Adicionar LIMIT apenas para prévia
                        if preview:
                            query += " LIMIT 50"
                        
                        query_result = conn.execute(text(query), {'date_format': date_format})
                    else:
                        return jsonify({'error': 'Filtro de data é obrigatório para consulta não consolidada'}), 400
            
            if query_result is None:
                return jsonify({'error': 'Nenhum dado encontrado'}), 404
            
            # Se for uma requisição de prévia, retornar dados em formato especial
            if preview:
                # Enviar colunas e dados separadamente
                columns = list(query_result.keys())
                rows = query_result.fetchall()
                
                # Converter dados para arrays (preserva ordem)
                data_rows = []
                for row in rows:
                    row_array = []
                    for i, value in enumerate(row):
                        row_array.append(value if value is not None else '')
                    data_rows.append(row_array)
                
                return jsonify({
                    'columns': columns,  # Array ordenado com nomes das colunas
                    'rows': data_rows,   # Array de arrays com os dados
                    'count': len(data_rows)
                }), 200
            
            # Para exportação Excel, usar mesma lógica
            rows = query_result.fetchall()
            if not rows:
                return jsonify({'error': 'Nenhum registro encontrado'}), 404
            
            # Obter nomes das colunas na ordem original
            column_names = list(query_result.keys())
            
            wb = Workbook()
            ws = wb.active
            sheet_name = 'Consolidado' if consolidated else date_filter.replace('-', '_')
            ws.title = sheet_name
            
            # Escrever cabeçalhos na ordem original
            for col_idx, header in enumerate(column_names, 1):
                ws.cell(row=1, column=col_idx, value=str(header))
            
            # Escrever dados
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    try:
                        safe_value = safe_excel_value(cell_value)
                        ws.cell(row=row_idx, column=col_idx, value=safe_value)
                    except Exception as cell_error:
                        ws.cell(row=row_idx, column=col_idx, value='')
            
            # Ajustar largura das colunas
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
            
            # Salvar Excel
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            period = 'consolidado' if consolidated else date_filter.replace('-', '')
            filename = f"{card_name}_{period}_{timestamp}.xlsx"
            
            return send_file(
                excel_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
    except Exception as e:
        print(f"Erro geral em export_card_data para {card_name}: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Erro ao exportar dados: {str(e)}'}), 500

# ========================================
# FUNÇÕES AUXILIARES
# ========================================

def format_response_data(data, decimal_fields=None):
    """Formata dados da resposta convertendo Decimal para float"""
    if decimal_fields is None:
        decimal_fields = []
    
    if isinstance(data, list):
        return [format_response_data(item, decimal_fields) for item in data]
    elif isinstance(data, dict):
        formatted = {}
        for key, value in data.items():
            if key in decimal_fields and isinstance(value, Decimal):
                formatted[key] = decimal_to_float(value)
            elif isinstance(value, (dict, list)):
                formatted[key] = format_response_data(value, decimal_fields)
            else:
                formatted[key] = decimal_to_float(value) if isinstance(value, Decimal) else value
        return formatted
    else:
        return decimal_to_float(data) if isinstance(value, Decimal) else data